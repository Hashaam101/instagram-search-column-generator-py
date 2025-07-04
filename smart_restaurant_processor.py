import pandas as pd
import phonenumbers
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.utils import get_column_letter
import os
import platform
from tqdm import tqdm
import unicodedata

# File paths
INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "output.xlsx"

# Column names (from original Excel headers)
NAME_COL = "Company name <COMPANY name>"
PHONE_COL = "Company Phone <Company Phone>"
LINK_COL = "Instagram_link"

# Global state
has_changed = False
instagram_found_count = 0
instagram_fallback_count = 0
duplicates_removed = 0
partial_duplicates = []
phone_format_count = 0
ran_actions = False

# Formatting
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
bold_font = Font(bold=True)


def clear_screen():
    if platform.system() == "Windows":
        os.system('cls')
    else:
        os.system('clear')


def load_data():
    try:
        df = pd.read_excel(INPUT_FILE, dtype=str)
        df.columns = [col.strip() for col in df.columns]
        df = df.dropna(how="all")
        return df
    except Exception as e:
        print(f"❌ Failed to load input file: {e}")
        exit(1)


def save_data(df):
    df = df.sort_values(by=NAME_COL, key=lambda x: x.str.lower())
    try:
        if os.path.exists(OUTPUT_FILE):
            os.remove(OUTPUT_FILE)

        df.to_excel(OUTPUT_FILE, index=False)
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        phone_col_idx = headers.index(PHONE_COL) + 1
        link_col_idx = headers.index(LINK_COL) + 1
        name_col_idx = headers.index(NAME_COL) + 1

        # Format phone column as text
        for row in range(2, ws.max_row + 1):
            phone_cell = ws.cell(row=row, column=phone_col_idx)
            phone_cell.number_format = FORMAT_TEXT

        # Convert Instagram links into HYPERLINK formulas (with visible URL)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=link_col_idx)
            if cell.value and not str(cell.value).startswith("=HYPERLINK"):
                cell.value = f'=HYPERLINK("{cell.value}", "{cell.value}")'

        # Highlight partial duplicates (red + bold)
        for dup in partial_duplicates:
            for row in range(2, ws.max_row + 1):
                name = str(ws.cell(row=row, column=name_col_idx).value).strip()
                phone = str(ws.cell(row=row, column=phone_col_idx).value).strip()
                if name == dup['name'] or phone == dup['phone']:
                    phone_cell = ws.cell(row=row, column=phone_col_idx)
                    phone_cell.font = bold_font
                    phone_cell.fill = red_fill

        # Highlight name duplicates (orange) if not already in red
        seen_names = {}
        for row in range(2, ws.max_row + 1):
            name = str(ws.cell(row=row, column=name_col_idx).value).strip().lower()
            if name in seen_names:
                for r in [seen_names[name], row]:
                    name_cell = ws.cell(row=r, column=name_col_idx)
                    if name_cell.fill != red_fill:
                        name_cell.fill = orange_fill
            else:
                seen_names[name] = row

        wb.save(OUTPUT_FILE)
    except Exception as e:
        print(f"❌ Failed to write output file: {e}")
        exit(1)


def normalize_phone(df):
    global phone_format_count, has_changed

    def format_number(num):
        try:
            num = str(num).strip().replace(".0", "")
            digits = ''.join(filter(str.isdigit, num))
            if len(digits) == 10:
                digits = '1' + digits
            if not digits.startswith('1'):
                digits = '1' + digits
            parsed = phonenumbers.parse("+" + digits, 'US')
            if phonenumbers.is_valid_number(parsed):
                return phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.INTERNATIONAL)
        except:
            return num
        return num

    df[PHONE_COL] = df[PHONE_COL].astype(str).apply(lambda x: format_number(x))
    phone_format_count = len(df)
    has_changed = True
    print(f"✔ {phone_format_count} phone numbers formatted.")
    return df


def remove_duplicates(df):
    global duplicates_removed, partial_duplicates, has_changed

    def normalize_phone(num):
        try:
            num = str(num).strip().replace(".0", "")
            digits = ''.join(filter(str.isdigit, num))
            if len(digits) == 10:
                digits = '1' + digits
            if not digits.startswith('1'):
                digits = '1' + digits
            parsed = phonenumbers.parse("+" + digits, 'US')
            if phonenumbers.is_valid_number(parsed):
                return phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.INTERNATIONAL)
        except:
            return str(num)
        return str(num)

    def normalize_name(name):
        if not isinstance(name, str):
            return ""
        name = unicodedata.normalize("NFKD", name)
        name = name.lower().strip()
        name = name.replace("’", "'").replace("‘", "'").replace("`", "'")
        name = ''.join(c for c in name if not unicodedata.category(c).startswith("C"))
        return name

    df['_NormalizedPhone'] = df[PHONE_COL].apply(normalize_phone)
    df['_NormalizedName'] = df[NAME_COL].apply(normalize_name)

    exact_dups = df[df.duplicated(subset=['_NormalizedName', '_NormalizedPhone'], keep=False)]
    duplicates_removed = len(exact_dups)
    exact_keys = set(exact_dups[['_NormalizedName', '_NormalizedPhone']].apply(tuple, axis=1))

    name_dups = df[df.duplicated(['_NormalizedName'], keep=False)]
    phone_dups = df[df.duplicated(['_NormalizedPhone'], keep=False)]

    seen = set()
    for _, row in pd.concat([name_dups, phone_dups]).iterrows():
        key = (row['_NormalizedName'], row['_NormalizedPhone'])
        if key not in seen and key not in exact_keys:
            partial_duplicates.append({'name': row[NAME_COL], 'phone': row[PHONE_COL]})
            seen.add(key)

    df = df.drop_duplicates(subset=['_NormalizedName', '_NormalizedPhone'])
    df.drop(columns=['_NormalizedName', '_NormalizedPhone'], inplace=True)

    has_changed = True
    print(f"✔ {duplicates_removed} exact duplicates removed.")
    print(f"⚠ {len(partial_duplicates)} partial duplicates found and marked.")
    return df

def generate_instagram_links(df):
    global instagram_fallback_count, has_changed

    print("🔍 Generating Instagram search links...")
    if LINK_COL not in df.columns:
        df[LINK_COL] = ""

    links = []
    for name in tqdm(df[NAME_COL], desc="Generating Instagram Search Links"):
        query = '+'.join(name.split())
        link = f"https://www.google.com/search?q={query}+restaurant+hawaii+site:instagram.com"
        links.append(link)
        instagram_fallback_count += 1

    df[LINK_COL] = links
    has_changed = True
    print(f"🔗 {instagram_fallback_count} Instagram search links generated.")
    return df


def main():
    global has_changed, ran_actions
    clear_screen()
    df = load_data()

    while True:
        print("\nWhat would you like to do?")
        print("[Press Enter] Run all smart functions")
        print("[1] Generate Instagram links only")
        print("[2] Remove duplicates only")
        print("[3] Improve phone number formatting only")
        print("[0] Export results to output.xlsx (requires at least one action)")

        choice = input("> ").strip()

        if choice == "":
            df = remove_duplicates(df)
            df = generate_instagram_links(df)
            df = normalize_phone(df)
            ran_actions = True
        elif choice == "1":
            df = generate_instagram_links(df)
            ran_actions = True
        elif choice == "2":
            df = remove_duplicates(df)
            ran_actions = True
        elif choice == "3":
            df = normalize_phone(df)
            ran_actions = True
        elif choice == "0":
            if has_changed or ran_actions:
                save_data(df)
                print(f"✔ Output written to {OUTPUT_FILE}")
                break
            else:
                print("⚠ You must perform at least one action before exporting.")
        else:
            print("Invalid option. Try again.")


if __name__ == '__main__':
    main()
